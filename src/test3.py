import sys
from functools import partial
from multiprocessing import Pool
from shared.common_utils import *
import open3d as o3d
import numpy as np
import openpyxl

import bs4
import requests

if __name__ == '__main__':
    # Read xls file
    xls_file = r'C:\Users\whats\Dropbox\Project\2023-发明报奖\广东省技术发明奖申报\专利检索报告\整理版.xlsx'
    xls_data = openpyxl.load_workbook(xls_file)
    xls_data = xls_data['下载的著录项1']
    for i in range(2, xls_data.max_row):
        if xls_data.cell(row=i, column=10).value == "发明申请":
            continue
        print(xls_data.cell(row=i, column=14).hyperlink.target)
        # print("{}: {}".format(
        #     xls_data.cell(row=i, column=2).value,
        #     xls_data.cell(row=i, column=14).hyperlink.target
        # ))

    pass
